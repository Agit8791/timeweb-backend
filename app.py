from flask import Flask, render_template, request, jsonify, send_file
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import random

app = Flask(__name__)

# In-memory storage
timetables = {}
# Short-term session memory: keyed by session_id, resets with process
session_memory = {}

class TimetableGenerator:
    def __init__(self, teachers, subjects, classrooms, time_slots, days, semesters):
        self.teachers = teachers
        self.subjects = subjects
        self.classrooms = classrooms
        self.time_slots = time_slots
        self.days = days
        self.semesters = semesters
        self.timetable = []
        self.conflicts = []
        

        
    def _subject_color(self, name: str) -> str:
        # Deterministic HSL pastel color from subject name
        h = 0
        for ch in name:
            h = (h * 31 + ord(ch)) % 360
        return f"hsl({h},65%,85%)"

    def _rank_slots(self, cohort_load, day, slot, cohort):
        # Lower score is better
        day_load = cohort_load.get(cohort, {}).get(day, 0)
        # prefer days with lower load and contiguous slots within a day
        contiguous_bonus = 0
        # in absence of exact time order semantics, no adjacency detection beyond load
        score = day_load * 10 - contiguous_bonus
        return score

    def generate(self):
        """Generate timetable with deterministic, exhaustive scheduling and validation for large datasets"""
        self.timetable = []
        self.conflicts = []

        # Pre-index teachers by subject (case-insensitive) for O(1) lookups
        subject_to_teachers = {}
        for t in self.teachers:
            for s in (t.get('subjects') or []):
                key = (s or '').strip().lower()
                if not key:
                    continue
                subject_to_teachers.setdefault(key, []).append(t)

        # Normalize day/slot ordering deterministically
        days_order = list(self.days)
        slots_order = list(self.time_slots)

        # Initialize schedules for tracking
        teacher_schedule = {
            teacher['name']: {day: {slot: False for slot in slots_order} for day in days_order}
            for teacher in self.teachers
        }
        classroom_schedule = {
            classroom: {day: {slot: False for slot in slots_order} for day in days_order}
            for classroom in self.classrooms
        }
        # cohort schedule by semester
        cohort_schedule = {
            semester: {day: {slot: False for slot in slots_order} for day in days_order}
            for semester in (set([s.get('semester', 'General') for s in self.subjects]) or {'General'})
        }
        # cohort daily load counts
        cohort_load = {}

        def can_place(subject, teacher, classroom, day, slot):
            semester = subject.get('semester', 'General')
            if teacher_schedule.get(teacher['name'], {}).get(day, {}).get(slot):
                return False
            if classroom_schedule.get(classroom, {}).get(day, {}).get(slot):
                return False
            if cohort_schedule.get(semester, {}).get(day, {}).get(slot):
                return False
            if not self.is_teacher_available(teacher, day, slot):
                return False
            return True

        def extract_all_dept_codes(subject):
            try:
                depts = subject.get('departments') or []
                if not isinstance(depts, list) or not depts:
                    return []
                codes = []
                for d in depts:
                    raw = str(d).strip()
                    if not raw:
                        continue
                    code = None
                    for sep in ['–', '—', '-']:
                        if sep in raw:
                            token = raw.split(sep, 1)[0].strip()
                            code = token if token else raw.split(sep, 1)[0].strip()
                            break
                    if code is None:
                        code = raw.split()[0]
                    if code:
                        codes.append(code)
                seen = set()
                uniq = []
                for c in codes:
                    if c not in seen:
                        seen.add(c)
                        uniq.append(c)
                return uniq
            except Exception:
                return []

        def place_entry(subject, teacher, classroom, day, slot):
            semester = subject.get('semester', 'General')
            entry = {
                'day': day,
                'time_slot': slot,
                'subject': subject['name'],
                'teacher': teacher['name'],
                'semester': semester,
                'classrooms': [classroom],
                'department_codes': extract_all_dept_codes(subject)
            }
            self.timetable.append(entry)
            teacher_schedule[teacher['name']][day][slot] = True
            classroom_schedule[classroom][day][slot] = True
            cohort_schedule[semester][day][slot] = True
            cohort_load.setdefault(semester, {}).setdefault(day, 0)
            cohort_load[semester][day] += 1

        def rank_slots_for_semester(semester):
            ranked = []
            for day in days_order:
                load = cohort_load.get(semester, {}).get(day, 0)
                for slot in slots_order:
                    score = self._rank_slots(cohort_load, day, slot, semester)
                    # include load primarily to spread out
                    ranked.append((score + load * 10, day, slot))
            ranked.sort(key=lambda x: (x[0], days_order.index(x[1]), slots_order.index(x[2])))
            return [(day, slot) for _, day, slot in ranked]

        def all_candidate_assignments(subject):
            """Yield deterministic feasible combos (day, slot, teacher, classroom) sorted by rank."""
            semester = subject.get('semester', 'General')
            subject_key = (subject.get('name') or '').strip().lower()
            teachers_for_subject = subject_to_teachers.get(subject_key, [])
            if not teachers_for_subject:
                return []
            candidates = []
            for day, slot in rank_slots_for_semester(semester):
                for t in teachers_for_subject:
                    if not self.is_teacher_available(t, day, slot):
                        continue
                    for c in self.classrooms:
                        if can_place(subject, t, c, day, slot):
                            # prioritize balanced teacher load: fewer total slots scheduled earlier
                            t_load = sum(1 for d in days_order for s in slots_order if teacher_schedule[t['name']][d][s])
                            candidates.append((t_load, days_order.index(day), slots_order.index(slot), day, slot, t, c))
            candidates.sort(key=lambda x: (x[0], x[1], x[2], (x[5]['name']).lower(), x[6]))
            return [(day, slot, t, c) for _, _, _, day, slot, t, c in candidates]

        # Assign classes per subject exhaustively
        for subject in self.subjects:
            sessions_required = int(subject.get('sessions_per_week', 2) or 0)
            placed = 0
            # Enumerate all feasible assignments once
            feasible = all_candidate_assignments(subject)
            used_positions = set()  # (day, slot) used by this subject to spread across days
            used_days = set()       # days already used for this subject to encourage spread
            feasible_days = set(d for (d, _s, _t, _c) in feasible)

            # Greedy deterministic placement favoring spread and low teacher load
            for (day, slot, teacher, classroom) in feasible:
                if placed >= sessions_required:
                    break
                # Avoid multiple sessions of same subject in same (day, slot)
                pos_key = (day, slot)
                if pos_key in used_positions:
                    continue
                # If there are enough distinct feasible days to cover all sessions, avoid reusing a day
                if len(feasible_days) >= sessions_required and day in used_days:
                    continue
                if can_place(subject, teacher, classroom, day, slot):
                    place_entry(subject, teacher, classroom, day, slot)
                    used_positions.add(pos_key)
                    used_days.add(day)
                    placed += 1

            # If still missing, allow same (day, slot) but try different teachers/classrooms
            if placed < sessions_required:
                for (day, slot, teacher, classroom) in feasible:
                    if placed >= sessions_required:
                        break
                    if can_place(subject, teacher, classroom, day, slot):
                        place_entry(subject, teacher, classroom, day, slot)
                        placed += 1

            # Record conflicts with deterministic suggestions if any sessions couldn't be placed
            if placed < sessions_required:
                missing = sessions_required - placed

                # Build explicit reasons for unplaced sessions
                reasons = []
                subject_key = (subject.get('name') or '').strip().lower()
                teachers_for_subject = subject_to_teachers.get(subject_key, [])
                if not teachers_for_subject:
                    reasons.append('No teacher associated with subject')
                else:
                    # For each day/slot, check why not placeable
                    for day in days_order:
                        for slot in slots_order:
                            any_teacher_free = any(self.is_teacher_available(t, day, slot) and not teacher_schedule[t['name']][day][slot]
                                                   for t in teachers_for_subject)
                            any_class_free = any(not classroom_schedule[c][day][slot] for c in self.classrooms)
                            semester = subject.get('semester', 'General')
                            cohort_busy = cohort_schedule[semester][day][slot]
                            if not any_teacher_free:
                                reasons.append(f"No available teacher at {day} {slot}")
                            if not any_class_free:
                                reasons.append(f"No available classroom at {day} {slot}")
                            if cohort_busy:
                                reasons.append(f"Semester busy at {day} {slot}")
                    # Deduplicate reasons
                    seen = set()
                    reasons = [r for r in reasons if not (r in seen or seen.add(r))]

                # Suggestions (top few free positions regardless of teacher)
                suggestions = []
                for day in days_order:
                    for slot in slots_order:
                        semester = subject.get('semester', 'General')
                        if cohort_schedule[semester][day][slot]:
                            continue
                        # gather teachers and classrooms free
                        free_teachers = [t for t in teachers_for_subject if self.is_teacher_available(t, day, slot) and not teacher_schedule[t['name']][day][slot]]
                        free_classes = [c for c in self.classrooms if not classroom_schedule[c][day][slot]]
                        if free_teachers and free_classes:
                            suggestions.append(f"{day} @ {slot}")
                        if len(suggestions) >= 5:
                            break
                    if len(suggestions) >= 5:
                        break

                self.conflicts.append({
                    'type': 'student',
                    'semester': subject.get('semester', 'General'),
                    'time_slot': None,
                    'day': None,
                    'subjects': [subject['name']],
                    'missing_sessions': missing,
                    'suggestions': suggestions,
                    'reasons': reasons
                })

        # Detect conflicts after placement for safety
        self.detect_conflicts()

        # Post-generation validation: ensure each subject appears required number of times
        occurrences = {}
        for e in self.timetable:
            occurrences[(e['subject'], e.get('semester', 'General'))] = occurrences.get((e['subject'], e.get('semester', 'General')), 0) + 1
        for subject in self.subjects:
            sessions_required = int(subject.get('sessions_per_week', 2) or 0)
            key = (subject['name'], subject.get('semester', 'General'))
            count = occurrences.get(key, 0)
            if count < sessions_required:
                missing = sessions_required - count
                # Append validation conflict if not already present
                self.conflicts.append({
                    'type': 'student',
                    'semester': subject.get('semester', 'General'),
                    'time_slot': None,
                    'day': None,
                    'subjects': [subject['name']],
                    'missing_sessions': missing,
                    'validation': True
                })

        return {
            'timetable': self.timetable,
            'conflicts': self.conflicts
        }
    
    def is_teacher_available(self, teacher, day, time_slot):
        """Check if teacher is available at given time"""
        availability = teacher.get('availability', {})
        # If no availability info provided, treat as fully available (backward compatible)
        if not availability:
            return True
        # If availability is provided, strictly enforce available days and slots
        if day not in availability:
            return False
        return time_slot in availability[day]
    
    def detect_conflicts(self):
        """Detect scheduling conflicts across teacher/classroom/student and attach suggestions"""
        conflicts = []
        indexed = {}
        for e in self.timetable:
            key = (e['day'], e['time_slot'])
            indexed.setdefault(key, []).append(e)

        for (day, slot), entries in indexed.items():
            # teacher conflicts
            teacher_map = {}
            classroom_map = {}
            cohort_map = {}
            for e in entries:
                teacher_map.setdefault(e['teacher'], []).append(e)
                # explode classrooms list for conflict detection per room
                for c in (e.get('classrooms') or []):
                    classroom_map.setdefault(c, []).append(e)
                cohort_map.setdefault(e.get('semester','General'), []).append(e)
            for t, arr in teacher_map.items():
                if len(arr) > 1:
                    conflicts.append({
                        'type': 'teacher', 'teacher': t, 'day': day, 'time_slot': slot,
                        'subjects': [x['subject'] for x in arr]
                    })
            for c, arr in classroom_map.items():
                if len(arr) > 1:
                    conflicts.append({
                        'type': 'classroom', 'classroom': c, 'day': day, 'time_slot': slot,
                        'subjects': [x['subject'] for x in arr]
                    })
            for cohort, arr in cohort_map.items():
                if len(arr) > 1:
                    conflicts.append({
                        'type': 'student', 'semester': cohort, 'day': day, 'time_slot': slot,
                        'subjects': [x['subject'] for x in arr]
                    })
        self.conflicts.extend(conflicts)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/generate', methods=['POST'])
def generate_timetable():
    try:
        data = request.json
        # Optional existing session_id to continue a session
        existing_session_id = data.get('session_id')
        
        # Pull from memory if fields are missing and a session exists
        memory = session_memory.get(existing_session_id, {}) if existing_session_id else {}
        
        teachers = data.get('teachers', memory.get('teachers', []))
        subjects = data.get('subjects', memory.get('subjects', []))
        classrooms = data.get('classrooms', memory.get('classrooms', []))
        time_slots = data.get('timeSlots', memory.get('timeSlots', []))
        days = data.get('days', memory.get('days', []))
        semesters = data.get('semesters', memory.get('semesters', []))
        preferences = data.get('preferences', memory.get('preferences', {}))
        
        if not all([teachers, subjects, classrooms, time_slots, days]):
            return jsonify({'error': 'Missing required data'}), 400
        
        generator = TimetableGenerator(teachers, subjects, classrooms, time_slots, days, semesters)
        result = generator.generate()
        
        # Decide session_id: reuse if provided, else create new
        import secrets
        session_id = existing_session_id or secrets.token_hex(8)
        timetables[session_id] = {
            'timetable': result['timetable'],
            'conflicts': result['conflicts'],
            'metadata': {
                'classrooms': classrooms,
                'days': days,
                'time_slots': time_slots,
                'semesters': semesters
            }
        }
        
        # Update short-term memory for this session
        session_memory[session_id] = {
            'teachers': teachers,
            'subjects': subjects,
            'classrooms': classrooms,
            'timeSlots': time_slots,
            'days': days,
            'semesters': semesters,
            'preferences': preferences,
            'last_updated': datetime.now().isoformat()
        }
        
        # Normalize timetable entries to unified format
        def normalize_entry(e):
            return {
                'day': e.get('day'),
                'time_slot': e.get('time_slot'),
                'subject': e.get('subject'),
                'teacher': e.get('teacher'),
                'semester': e.get('semester'),
                'classrooms': e.get('classrooms') if isinstance(e.get('classrooms'), list) and e.get('classrooms') else ([e.get('classroom')] if e.get('classroom') else []),
                'department_codes': e.get('department_codes') if isinstance(e.get('department_codes'), list) else ([] if e.get('department_codes') is None else [e.get('department_codes')]),
                'description': e.get('description')
            }
        normalized_tt = [normalize_entry(e) for e in result['timetable']]

        return jsonify({
            'success': True,
            'session_id': session_id,
            'timetable': normalized_tt,
            'conflicts': result['conflicts'],
            'memory': session_memory[session_id]
        })
        
    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/update-timetable', methods=['POST'])
def update_timetable():
    try:
        data = request.json
        session_id = data.get('session_id')
        timetable = data.get('timetable')
        # Optional updates to memory fields to let user modify previous inputs
        mem_updates = data.get('memory_updates', {})

        # Normalize incoming timetable to unified format (arrays for classrooms/department_codes)
        def normalize_entry(e):
            classrooms = e.get('classrooms')
            if not isinstance(classrooms, list):
                classrooms = [e.get('classroom')] if e.get('classroom') else []
            dept = e.get('department_codes')
            if not isinstance(dept, list):
                dept = ([] if dept is None else [dept])
            return {
                'day': e.get('day'),
                'time_slot': e.get('time_slot'),
                'subject': e.get('subject'),
                'teacher': e.get('teacher'),
                'semester': e.get('semester'),
                'classrooms': classrooms,
                'department_codes': dept,
                'description': e.get('description')
            }
        normalized_tt = [normalize_entry(e) for e in (timetable or [])]
        
        if session_id in timetables:
            timetables[session_id]['timetable'] = normalized_tt
            
            # Re-detect conflicts
            generator = TimetableGenerator([], [], [], [], [], [])
            generator.timetable = normalized_tt
            generator.detect_conflicts()
            timetables[session_id]['conflicts'] = generator.conflicts

            # Apply memory updates if provided
            if session_id in session_memory and isinstance(mem_updates, dict):
                session_memory[session_id].update({k: v for k, v in mem_updates.items() if v is not None})
                session_memory[session_id]['last_updated'] = datetime.now().isoformat()
            
            return jsonify({
                'success': True,
                'conflicts': generator.conflicts,
                'memory': session_memory.get(session_id)
            })
        
        return jsonify({'error': 'Session not found'}), 404
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/export/excel/<session_id>')
def export_excel(session_id):
    try:
        if session_id not in timetables:
            return "Timetable not found", 404
        
        timetable_data = timetables[session_id]['timetable']
        metadata = timetables[session_id]['metadata']
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)
        
        classrooms = metadata['classrooms']
        days = metadata['days']
        time_slots = metadata['time_slots']
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        day_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        day_font = Font(bold=True, size=10)
        break_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        break_font = Font(bold=True, color="856404", size=10)
        class_fill = PatternFill(start_color="E7F3E7", end_color="E7F3E7", fill_type="solid")
        title_font = Font(bold=True, size=14, color="4472C4")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Create a single sheet combining all classrooms
        ws = wb.create_sheet(title='Unified Timetable')

        # Title
        ws['A1'] = 'Unified Weekly Timetable'
        ws['A1'].font = title_font
        # Columns: Day + for each time slot, one column containing concatenated classroom entries
        total_cols = 1 + len(time_slots)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # Column headers (time slots)
        ws['A2'] = 'Day / Time'
        ws['A2'].fill = header_fill
        ws['A2'].font = header_font
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A2'].border = thin_border

        for col, slot in enumerate(time_slots, start=2):
            cell = ws.cell(row=2, column=col)
            cell.value = slot
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        # Fill data: for each day and slot, list all classroom entries combined
        for row, day in enumerate(days, start=3):
            day_cell = ws.cell(row=row, column=1)
            day_cell.value = day
            day_cell.fill = day_fill
            day_cell.font = day_font
            day_cell.alignment = Alignment(horizontal='center', vertical='center')
            day_cell.border = thin_border

            for col, slot in enumerate(time_slots, start=2):
                # Collect entries for all classrooms at this day/slot
                entries = [
                    e for e in timetable_data
                    if e['day'] == day and e['time_slot'] == slot
                ]
                cell = ws.cell(row=row, column=col)
                if entries:
                    # Compose multiline text: Classroom(s): Subject (Semester) - Teacher
                    formatted_blocks = []
                    for e in entries:
                        cls_list = e.get('classrooms') or []
                        classroom_label = ', '.join(cls_list) if cls_list else '-'
                        block = f"{classroom_label}: {e['subject']}\n{e['teacher']} ({e.get('semester', '-')})"
                        if e.get('department_codes'):
                            block += f"\nDept: {', '.join(e.get('department_codes') or [])}"
                        if e.get('description'):
                            block += f"\n{e.get('description')}"
                        formatted_blocks.append(block)
                    cell.value = "\n\n".join(formatted_blocks)
                    cell.fill = class_fill
                    cell.font = Font(size=9)
                else:
                    cell.value = ""
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

        # Adjust dimensions
        ws.column_dimensions['A'].width = 15
        for col in range(2, len(time_slots) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 28

        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 30
        for row in range(3, len(days) + 3):
            ws.row_dimensions[row].height = 90
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'timetable_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return str(e), 500

# Memory utility endpoints for frontend integration
@app.route('/memory/<session_id>', methods=['GET'])
def get_memory(session_id):
    if session_id in session_memory:
        return jsonify({'success': True, 'memory': session_memory[session_id]})
    return jsonify({'error': 'Session not found'}), 404

@app.route('/memory/<session_id>', methods=['POST'])
def update_memory(session_id):
    data = request.json or {}
    if session_id not in session_memory:
        session_memory[session_id] = {}
    session_memory[session_id].update({k: v for k, v in data.items() if v is not None})
    session_memory[session_id]['last_updated'] = datetime.now().isoformat()
    return jsonify({'success': True, 'memory': session_memory[session_id]})

@app.route('/memory/clear/<session_id>', methods=['POST'])
def clear_memory(session_id):
    if session_id in session_memory:
        del session_memory[session_id]
        return jsonify({'success': True})
    return jsonify({'error': 'Session not found'}), 404

@app.route('/export/pdf/<session_id>')
def export_pdf(session_id):
    try:
        if session_id not in timetables:
            return "Timetable not found", 404
        
        timetable_data = timetables[session_id]['timetable']
        metadata = timetables[session_id]['metadata']
        conflicts = timetables[session_id].get('conflicts', [])
        
        # Create PDF
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output, 
            pagesize=landscape(A3), 
            topMargin=0.5*inch, 
            bottomMargin=0.5*inch,
            leftMargin=0.5*inch,
            rightMargin=0.5*inch
        )
        elements = []
        
        styles = getSampleStyleSheet()
        classrooms = metadata['classrooms']
        days = metadata['days']
        time_slots = metadata['time_slots']
        
        # Create a single unified timetable across all classrooms
        title = Paragraph("<b>Unified Weekly Timetable</b>", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.2 * inch))

        # Create table data: Day/Time + for each slot, one cell containing combined entries for all classrooms
        table_data = [['Day / Time'] + time_slots]

        for day in days:
            row = [day]
            for slot in time_slots:
                entries = [e for e in timetable_data if e['day'] == day and e['time_slot'] == slot]
                if entries:
                    # Build visually separated entries using themed bullet separators
                    themed_sep = " \u2022 "  # bullet dot separator
                    formatted = []
                    for e in entries:
                        cls_list = e.get('classrooms') or []
                        classroom_label = ', '.join(cls_list) if cls_list else '-'
                        block_lines = [
                            f"{classroom_label}: {e['subject']}",
                            f"{e['teacher']} ({e.get('semester', '-')})"
                        ]
                        if e.get('department_codes'):
                            block_lines.append(f"Dept: {', '.join(e.get('department_codes') or [])}")
                        if e.get('description'):
                            block_lines.append(e['description'])
                        formatted.append("\n".join(block_lines))
                    # Join blocks with a clear separator line and extra spacing for readability
                    row.append((f"\n{themed_sep}\n").join(formatted))
                else:
                    row.append('')
            table_data.append(row)

        # Compute column widths to span full page width
        available_width = doc.width
        n_time_cols = len(time_slots)
        day_col_width = max(72, available_width * 0.12)
        remaining_width = max(0, available_width - day_col_width)
        slot_col_width = remaining_width / n_time_cols if n_time_cols > 0 else remaining_width
        col_widths = [day_col_width] + [slot_col_width] * n_time_cols

        # Create and style table with enhanced design
        table = Table(table_data, repeatRows=1, colWidths=col_widths)

        header_bg = colors.Color(0.18, 0.36, 0.6)
        header_text = colors.whitesmoke
        day_col_bg = colors.HexColor('#E9EFF8')
        grid_color = colors.HexColor('#B0BEC5')
        zebra_bg = colors.HexColor('#F7FAFC')
        entry_sep_color = colors.HexColor('#DDE6F3')

        style_list = [
            ('BACKGROUND', (0, 0), (-1, 0), header_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), header_text),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (0, -1), day_col_bg),
            ('TEXTCOLOR', (0, 1), (0, -1), colors.HexColor('#0F3057')),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (0, -1), 10),
            ('FONTNAME', (1, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (1, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, grid_color),
        ]

        # Zebra striping for rows (excluding header)
        for r in range(1, len(table_data)):
            if r % 2 == 0:
                style_list.append(('BACKGROUND', (1, r), (-1, r), zebra_bg))

        # Per-cell subject color is not trivial in a combined multi-entry cell; keep readable base and subtle highlight if any entries exist
        for row_idx, day in enumerate(days, start=1):
            for col_idx, slot in enumerate(time_slots, start=1):
                has_any = any(e for e in timetable_data if e['day'] == day and e['time_slot'] == slot)
                if has_any:
                    style_list.append(('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#E8F5E9')))

        # Add subtle inner separators inside cells that have multiple entries
        for r, day in enumerate(days, start=1):
            for c, slot in enumerate(time_slots, start=1):
                es = [e for e in timetable_data if e['day'] == day and e['time_slot'] == slot]
                if len(es) > 1:
                    # Slight background tint to highlight multi-entry cells already set above; keep additional separator borders light
                    style_list.append(('LINEBEFORE', (c, r), (c, r), 0, entry_sep_color))
                    style_list.append(('LINEAFTER', (c, r), (c, r), 0, entry_sep_color))
        table.setStyle(TableStyle(style_list))
        elements.append(table)

        # Conflicts page
        if conflicts:
            elements.append(PageBreak())
            elements.append(Paragraph("<b>Conflicts</b>", styles['Title']))
            elements.append(Spacer(1, 0.2*inch))

            conflict_headers = ['Type','Day','Time','Teacher','Classroom','Semester','Subjects','Suggestions']
            conflict_rows = [conflict_headers]
            type_map = {'teacher':'Teacher','classroom':'Classroom','student':'Student'}

            for c in conflicts:
                conflict_rows.append([
                    type_map.get(c.get('type'), 'Conflict'),
                    c.get('day','-') or '-',
                    c.get('time_slot','-') or '-',
                    c.get('teacher','-') or '-',
                    c.get('classroom','-') or '-',
                    c.get('semester','-') or '-',
                    ", ".join(c.get('subjects', []) or []),
                    ", ".join(c.get('suggestions', []) or [])
                ])
            conflicts_table = Table(conflict_rows, repeatRows=1, colWidths=[80,60,60,110,110,90,200,200])
            conflicts_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), header_bg),
                ('TEXTCOLOR', (0,0), (-1,0), header_text),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('GRID', (0,0), (-1,-1), 0.5, grid_color),
                ('FONTSIZE', (0,0), (-1,0), 11),
                ('FONTSIZE', (0,1), (-1,-1), 8),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(conflicts_table)
        
        doc.build(elements)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'timetable_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return str(e), 500

@app.route('/conflicts/<session_id>', methods=['GET'])
def get_conflicts(session_id):
    if session_id not in timetables:
        return jsonify({'error': 'Session not found'}), 404
    return jsonify({'success': True, 'conflicts': timetables[session_id].get('conflicts', [])})

if __name__ == '__main__':
    app.run(debug=True, port=5000)